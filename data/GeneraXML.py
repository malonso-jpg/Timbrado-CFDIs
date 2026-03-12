#!/usr/bin/env python3
"""
Script para generar XMLs de nómina desde un archivo Excel.
Genera un XML por cada fila (empleado) del Excel.
Versión 3.0 - Con nuevo formato de columnas que incluye claves:
  P_{TipoPercepcion}_{Clave}_{GRAVADO/EXENTO}_{Descripcion}
  D_{TipoDeduccion}_{Clave}_{Descripcion}
  OP_{TipoOtroPago}_{Clave}_{Descripcion}
"""

import openpyxl
import xml.etree.ElementTree as ET
from datetime import date, datetime
from decimal import Decimal
import os
import json
from decimal import Decimal, ROUND_HALF_UP
import uuid
from data.funcionesPostgres import conexion
import traceback

# Registrar namespaces para el XML del SAT
ET.register_namespace('cfdi', 'http://www.sat.gob.mx/cfd/4')
ET.register_namespace('nomina12', 'http://www.sat.gob.mx/nomina12')
ET.register_namespace('xsi', 'http://www.w3.org/2001/XMLSchema-instance')

class ConfiguracionCampos:
    """Configuración de campos que nunca se incluirán aunque tengan valor"""
    
    # Campos que NUNCA se incluirán en el Receptor de Nómina
    CAMPOS_EXCLUIDOS_RECEPTOR_NOMINA = [
        # Descomenta los campos que NO quieres incluir:
        # 'NumSeguridadSocial',
        # 'FechaInicioRelLaboral',
        # 'Antiguedad',
        # 'Departamento',
        # 'Puesto',
        # 'RiesgoPuesto',
        # 'Banco',
        # 'CuentaBancaria',
        # 'SalarioBaseCotApor',
        # 'SalarioDiarioIntegrado',
    ]
    
    # Campos que NUNCA se incluirán en general
    CAMPOS_EXCLUIDOS_GENERALES = [
        # Agrega aquí otros campos que no quieras incluir
    ]
    
    @classmethod
    def campo_excluido(cls, campo):
        """Verifica si un campo está en la lista de exclusión"""
        return campo in cls.CAMPOS_EXCLUIDOS_RECEPTOR_NOMINA or campo in cls.CAMPOS_EXCLUIDOS_GENERALES
    
    @classmethod
    def guardar_configuracion(cls, filepath):
        """Guarda la configuración actual en un archivo JSON"""
        config = {
            'campos_excluidos_receptor_nomina': cls.CAMPOS_EXCLUIDOS_RECEPTOR_NOMINA,
            'campos_excluidos_generales': cls.CAMPOS_EXCLUIDOS_GENERALES
        }
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(config, f, indent=2, ensure_ascii=False)
    
    @classmethod
    def cargar_configuracion(cls, filepath):
        """Carga la configuración desde un archivo JSON"""
        if os.path.exists(filepath):
            with open(filepath, 'r', encoding='utf-8') as f:
                config = json.load(f)
                cls.CAMPOS_EXCLUIDOS_RECEPTOR_NOMINA = config.get('campos_excluidos_receptor_nomina', [])
                cls.CAMPOS_EXCLUIDOS_GENERALES = config.get('campos_excluidos_generales', [])


class ExcelToNominaXML:
    def __init__(self, excel_path, output_dir, config_path=None, ambiente="TEST"):
        self.excel_path = excel_path
        self.output_dir = output_dir
        self.reportes = []
        self.ambiente = ambiente.upper() if ambiente.upper() in ("TEST", "PROD") else "TEST"
        # Cargar configuración si existe
        if config_path and os.path.exists(config_path):
            ConfiguracionCampos.cargar_configuracion(config_path)
        
        # Valores por defecto para campos CFDI
        self.defaults = {
            'lugar_expedicion': '03200',
            'metodo_pago': 'PUE',
            'moneda': 'MXN',
            'tipo_comprobante': 'N',
            'exportacion': '01',
            'version': '4.0',
            'envio_cfdi_email': 'smorales@intelitax.com',
            'envio_cfdi_zip': '1',
            'envio_cfdi_pdf': '1',
            'envio_cfdi_xml': '1'
        }
    

    def decimal_default(self, obj):
        if isinstance(obj, Decimal):
            return float(obj)
        raise TypeError
    def _Deci(self,x):
        return Decimal(str(x or 0))
    def money_fmt(self, value, decimals=2):
        """
        Convierte cualquier número a string con N decimales
        usando redondeo fiscal (ROUND_HALF_UP).
        Devuelve '0.00' si viene None.
        """
        if value is None or value == "":
            value = Decimal("0")
        else:
            value = Decimal(str(value))

        q = Decimal("1." + "0" * decimals)  # 1.00
        return str(value.quantize(q, rounding=ROUND_HALF_UP))
    def _to_date(self, v):
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        if isinstance(v, str):
            v = v.strip()
            for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y"):
                try:
                    return datetime.strptime(v, fmt).date()
                except:
                    pass
        return None

    def antiguedad_pac(self, fecha_inicio, fecha_final):
        fi = self._to_date(fecha_inicio)
        ff = self._to_date(fecha_final)
        if not fi or not ff:
            return None
        if ff < fi:
            fi, ff = ff, fi
        dias = (ff - fi).days
        semanas = max(1, dias // 7)   # mínimo 1 para evitar P0W
        return f"P{semanas}W"
    def normalizar_cp(self,cp):
        if cp is None:
            return None
        return str(cp).strip().zfill(5)
    def ceor_izquierda(self,valor):
        if valor is None:
            return None
        return str(valor).strip().zfill(2)
    def tiene_valor(self, valor):
        """Verifica si un valor es válido (no None, no vacío, no 0)"""
        if valor is None:
            return False
        if isinstance(valor, str) and valor.strip() == '':
            return False
        if isinstance(valor, (int, float, Decimal)) and valor == 0:
            return False
        return True
    
    def limpiar_texto(self, texto):
        """Limpia un texto: trim, elimina espacios múltiples y retorna string limpio"""
        if texto is None:
            return ''
        
        texto_str = str(texto)
        # Eliminar espacios al inicio y final
        texto_str = texto_str.strip()
        # Reemplazar múltiples espacios por uno solo
        import re
        texto_str = re.sub(r'\s+', ' ', texto_str)
        
        return texto_str
    
    def leer_excel(self):
        """Lee el archivo Excel y devuelve los datos estructurados"""
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        
        # Obtener headers (fila 1)
        headers = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                headers.append(str(cell.value))
            else:
                headers.append(None)
        
        # Leer todas las filas de datos (desde fila 2)
        empleados = []
        for row_idx in range(2, ws.max_row + 1):
            empleado = {}
            for col_idx, header in enumerate(headers, 1):
                if header:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    empleado[header] = cell.value
            
            # Solo agregar si tiene datos (al menos RFC_receptor)
            if empleado.get('RFC_receptor'):
                empleado["__excel_row__"] = row_idx 
                empleados.append(empleado)
        
        wb.close()
        return empleados
    
    def extraer_percepciones(self, empleado):
        """Extrae y agrupa percepciones del empleado"""
        percepciones = {}
        
        for key, value in empleado.items():
            if key and key.startswith('P_') and self.tiene_valor(value):
                # Formato NUEVO: P_001_001_GRAVADO_SUELDOS_SALARIOS_RAYAS_Y_JORNALES
                # Formato: P_{TipoPercepcion}_{Clave}_{GRAVADO/EXENTO}_{Descripcion}
                parts = key.split('_', 4)
                if len(parts) >= 4:
                    tipo = parts[1]  # 001, 019, etc.
                    clave = parts[2]  # 001, 002, etc. (ahora viene del Excel)
                    gravado_exento = parts[3]  # GRAVADO o EXENTO
                    concepto = parts[4] if len(parts) > 4 else ''
                    
                    # Crear clave única por tipo de percepción
                    if tipo not in percepciones:
                        percepciones[tipo] = {
                            'tipo': tipo,
                            'clave': clave,  # NUEVO: guardar la clave
                            'concepto': concepto.replace('_', ' '),
                            'gravado': Decimal('0'),
                            'exento': Decimal('0')
                        }
                    
                    # Agregar importe
                    if gravado_exento == 'GRAVADO':
                        percepciones[tipo]['gravado'] = Decimal(str(value))
                    elif gravado_exento == 'EXENTO':
                        percepciones[tipo]['exento'] = Decimal(str(value))
        
        return list(percepciones.values())
    
    def extraer_deducciones(self, empleado):
        """Extrae deducciones del empleado"""
        deducciones = []
        
        for key, value in empleado.items():
            if key and key.startswith('D_') and self.tiene_valor(value):
                # Formato NUEVO: D_001_001_IMSS
                # Formato: D_{TipoDeduccion}_{Clave}_{Descripcion}
                parts = key.split('_', 3)
                if len(parts) >= 3:
                    tipo = parts[1]  # 001, 002, 003, etc.
                    clave = parts[2]  # 001, 002, etc. (ahora viene del Excel)
                    concepto = parts[3] if len(parts) > 3 else ''
                    
                    deducciones.append({
                        'tipo': tipo,
                        'clave': clave,  # NUEVO: guardar la clave
                        'concepto': concepto.replace('_', ' '),
                        'importe': Decimal(str(value))
                    })
        
        return deducciones
    
    def extraer_otros_pagos(self, empleado):
        """Extrae otros pagos del empleado"""
        otros_pagos = []
        
        for key, value in empleado.items():
            if key and key.startswith('OP_') and self.tiene_valor(value):
                # Formato NUEVO: OP_002_002_SUBSIDIO_CAUSADO
                # Formato: OP_{TipoOtroPago}_{Clave}_{Descripcion}
                parts = key.split('_', 3)
                if len(parts) >= 3:
                    tipo = parts[1]  # 002, 999, etc.
                    clave = parts[2]  # 002, 006, etc. (ahora viene del Excel)
                    concepto = parts[3] if len(parts) > 3 else ''
                    
                    otros_pagos.append({
                        'tipo': tipo,
                        'clave': clave,  # NUEVO: guardar la clave
                        'concepto': concepto.replace('_', ' '),
                        'importe': Decimal(str(value))
                    })
        
        return otros_pagos
    
    def calcular_totales(self, percepciones, deducciones, otros_pagos):
        """Calcula los totales"""
        calc = {}
        # Percepciones
        total_gravado = sum(self._Deci(p.get('gravado')) for p in percepciones)
        total_exento  = sum(self._Deci(p.get('exento'))  for p in percepciones)
        total_sueldos = total_gravado + total_exento  # (TotalSueldos)

        # Deducciones (ISR = tipo 002)
        impuestos_retenidos = sum(self._Deci(d.get('importe')) for d in deducciones if str(d.get('tipo')) == '002')
        otras_deducciones   = sum(self._Deci(d.get('importe')) for d in deducciones if str(d.get('tipo')) != '002')
        total_deducciones   = impuestos_retenidos + otras_deducciones
        
        # Otros pagos
        total_otros_pagos = sum(self._Deci(op.get('importe')) for op in otros_pagos)
        
        # CFDI Concepto "Pago de nómina"
        importe_conceptos   = total_sueldos + total_otros_pagos
        descuento_conceptos = total_deducciones
        subtotal            = importe_conceptos
        total               = subtotal - descuento_conceptos
        calc['valor_unitario']       = importe_conceptos
        calc['importe_conceptos']    = importe_conceptos
        calc['descuento_conceptos']  = descuento_conceptos
        calc['subtotal']             = subtotal
        calc['total']                = total
        # Totales Nómina
        calc['total_percepciones']          = total_sueldos
        calc['total_deducciones']           = total_deducciones
        calc['total_otros_pagos']           = total_otros_pagos
        calc['total_gravado']               = total_gravado
        calc['total_exento']                = total_exento
        calc['total_sueldos']               = total_sueldos
        calc['total_otras_deducciones']     = otras_deducciones
        calc['total_impuestos_retenidos']   = impuestos_retenidos
        
        return calc
    
    def formatear_fecha(self, fecha_value):
        """Convierte fecha de Excel al formato requerido"""
        if isinstance(fecha_value, datetime):
            return fecha_value.strftime('%Y-%m-%dT%H:%M:%S')
        elif isinstance(fecha_value, str):
            try:
                dt = datetime.strptime(fecha_value, '%Y-%m-%d %H:%M:%S')
                return dt.strftime('%Y-%m-%dT%H:%M:%S')
            except:
                try:
                    dt = datetime.strptime(fecha_value, '%Y-%m-%d')
                    return dt.strftime('%Y-%m-%dT%H:%M:%S')
                except:
                    return fecha_value
        return str(fecha_value)
    
    def formatear_fecha_simple(self, fecha_value):
        """Convierte fecha de Excel al formato YYYY-MM-DD"""
        if isinstance(fecha_value, datetime):
            return fecha_value.strftime('%Y-%m-%d')
        elif isinstance(fecha_value, str):
            try:
                dt = datetime.strptime(fecha_value, '%Y-%m-%d %H:%M:%S')
                return dt.strftime('%Y-%m-%d')
            except:
                try:
                    dt = datetime.strptime(fecha_value, '%Y-%m-%d')
                    return dt.strftime('%Y-%m-%d')
                except:
                    return fecha_value
        return str(fecha_value)
    
    def generar_xml(self, empleado):
        """Genera el XML de nómina para un empleado según Anexo 20 del SAT"""
        
        # Extraer percepciones, deducciones y otros pagos
        percepciones = self.extraer_percepciones(empleado)
        deducciones = self.extraer_deducciones(empleado)
        otros_pagos = self.extraer_otros_pagos(empleado)
        # Calcular totales
        totales_calc = self.calcular_totales(percepciones, deducciones, otros_pagos)
        # Crear reporte de diferencias
        reporte = self.crear_reporte_diferencias(empleado, totales_calc)
        self.reportes.append(reporte)
        
        # Crear documento raíz con namespaces (sin duplicar)
        comprobante = ET.Element('{http://www.sat.gob.mx/cfd/4}Comprobante')
        
        # Agregar namespaces
        comprobante.set('{http://www.w3.org/2001/XMLSchema-instance}schemaLocation',
            'http://www.sat.gob.mx/cfd/4 http://www.sat.gob.mx/sitio_internet/cfd/4/cfdv40.xsd '
            'http://www.sat.gob.mx/nomina12 http://www.sat.gob.mx/sitio_internet/cfd/nomina/nomina12.xsd')
        
        # Atributos del Comprobante según Anexo 20
        version_cfdi = empleado.get('Version_CFDI', 4)
        comprobante.set('Version', str(version_cfdi) if '.' in str(version_cfdi) else f"{version_cfdi}.0")
        
        if self.tiene_valor(empleado.get('Serie')):
            comprobante.set('Serie', str(empleado.get('Serie')))
        
        if self.tiene_valor(empleado.get('Folio')):
            comprobante.set('Folio', str(empleado.get('Folio')))
        
        comprobante.set('Fecha', self.formatear_fecha(empleado.get('Fecha', datetime.now())))
        
        # SubTotal ###mati: ahora el subtotal es la suma de percepciones + otros pagos, sin restar deducciones
        comprobante.set('SubTotal', self.money_fmt(totales_calc['subtotal']))
        if(totales_calc['descuento_conceptos'] > 0):
            comprobante.set('Descuento', self.money_fmt(totales_calc['descuento_conceptos']))
        # Moneda
        comprobante.set('Moneda', str(empleado.get('Moneda', self.defaults['moneda'])))
        
        # Total
        comprobante.set('Total', self.money_fmt(totales_calc['total']))
        
        # TipoDeComprobante
        comprobante.set('TipoDeComprobante', str(empleado.get('TipoDeComprobante', self.defaults['tipo_comprobante'])))
        
        # Exportacion
        comprobante.set('Exportacion', str(empleado.get('Exportacion', self.defaults['exportacion'])))
        
        # MetodoPago
        comprobante.set('MetodoPago', str(empleado.get('MetodoPago', self.defaults['metodo_pago'])))
        
        # LugarExpedicion
        lugar_expedicion = empleado.get('LugarExpedicion')
        if not self.tiene_valor(lugar_expedicion):
            lugar_expedicion = self.defaults['lugar_expedicion']
        comprobante.set('LugarExpedicion', str(lugar_expedicion))
        
        # CfdiRelacionados - solo si tiene TipoRelacion y UUID
        tipo_relacion = empleado.get('Tipo de relaciòn')
        uuid_relacionado = empleado.get('UUID')
        if self.tiene_valor(tipo_relacion) and self.tiene_valor(uuid_relacionado):
            cfdi_relacionados = ET.SubElement(comprobante, '{http://www.sat.gob.mx/cfd/4}CfdiRelacionados')
            tipo_rel_str = str(tipo_relacion).zfill(2) if len(str(tipo_relacion)) == 1 else str(tipo_relacion)
            cfdi_relacionados.set('TipoRelacion', tipo_rel_str)
            
            cfdi_relacionado = ET.SubElement(cfdi_relacionados, '{http://www.sat.gob.mx/cfd/4}CfdiRelacionado')
            cfdi_relacionado.set('UUID', str(uuid_relacionado))
        
        # Emisor
        emisor = ET.SubElement(comprobante, '{http://www.sat.gob.mx/cfd/4}Emisor')
        emisor.set('Rfc', self.limpiar_texto(empleado.get('RFC_emisor', '')))
        emisor.set('Nombre', self.limpiar_texto(empleado.get('Razon_emisor', '')))
        emisor.set('RegimenFiscal', self.limpiar_texto(empleado.get('Regimen_emisor', '')))
        
        # Receptor
        receptor = ET.SubElement(comprobante, '{http://www.sat.gob.mx/cfd/4}Receptor')
        receptor.set('Rfc', self.limpiar_texto(empleado.get('RFC_receptor', '')))
        receptor.set('Nombre', self.limpiar_texto(empleado.get('Razon_receptor', '')))
        receptor.set('DomicilioFiscalReceptor', self.normalizar_cp(self.limpiar_texto(empleado.get('Domicilio_receptor', ''))))
        receptor.set('RegimenFiscalReceptor', self.limpiar_texto(empleado.get('Regimen_receptor', '')))
        receptor.set('UsoCFDI', self.limpiar_texto(empleado.get('Uso_CFDI', '')))
        
        # Conceptos
        conceptos_elem = ET.SubElement(comprobante, '{http://www.sat.gob.mx/cfd/4}Conceptos')
        concepto_elem = ET.SubElement(conceptos_elem, '{http://www.sat.gob.mx/cfd/4}Concepto')
        concepto_elem.set('ClaveProdServ', self.limpiar_texto(empleado.get('ClaveProdServ', '84111505')))
        concepto_elem.set('Cantidad', str(empleado.get('Cantidad', '1')))
        concepto_elem.set('ClaveUnidad', self.limpiar_texto(empleado.get('ClaveUnidad', 'ACT')))
        concepto_elem.set('Descripcion', self.limpiar_texto(empleado.get('Descripcion', 'Pago de nómina')))

        ###mati
        concepto_elem.set('ValorUnitario', self.money_fmt(totales_calc['valor_unitario']))
        concepto_elem.set('Importe', self.money_fmt(totales_calc['importe_conceptos']))
        if(totales_calc['descuento_conceptos'] > 0):
            concepto_elem.set('Descuento', self.money_fmt(totales_calc['descuento_conceptos']))
        #

        concepto_elem.set('ObjetoImp', self.limpiar_texto(empleado.get('ObjetoImp', '01')))
        
        # Complemento
        complemento = ET.SubElement(comprobante, '{http://www.sat.gob.mx/cfd/4}Complemento')
        
        # Nómina
        nomina = ET.SubElement(complemento, '{http://www.sat.gob.mx/nomina12}Nomina')
        nomina.set('Version', str(empleado.get('Versiòn', '1.2')))
        nomina.set('TipoNomina', str(empleado.get('Tipo_nomina', '')))
        nomina.set('FechaPago', self.formatear_fecha_simple(empleado.get('FechaPago', '')))
        nomina.set('FechaInicialPago', self.formatear_fecha_simple(empleado.get('Fecha_inicial_pago', '')))
        nomina.set('FechaFinalPago', self.formatear_fecha_simple(empleado.get('Fecha_final_pago', '')))
        
        # Formatear días pagados
        dias_pagados = empleado.get('Dias_pagados', 0)
        if isinstance(dias_pagados, (int, float)):
            nomina.set('NumDiasPagados', f"{dias_pagados:.3f}")
        else:
            nomina.set('NumDiasPagados', str(dias_pagados))
        
        # TotalPercepciones (solo si hay percepciones)
        if totales_calc['total_percepciones'] > 0:
            nomina.set('TotalPercepciones', f"{float(totales_calc['total_percepciones']):.2f}")
        
        # TotalDeducciones (solo si hay deducciones)
        if totales_calc['total_deducciones'] > 0:
            nomina.set('TotalDeducciones', f"{float(totales_calc['total_deducciones']):.2f}")
        
        # TotalOtrosPagos (solo si hay otros pagos)
        if totales_calc['total_otros_pagos'] > 0:
            nomina.set('TotalOtrosPagos', f"{float(totales_calc['total_otros_pagos']):.2f}")
        
        # Emisor nómina
        emisor_nom = ET.SubElement(nomina, '{http://www.sat.gob.mx/nomina12}Emisor')
        if self.tiene_valor(empleado.get('Registro_patronal')):
            emisor_nom.set('RegistroPatronal', self.limpiar_texto(empleado.get('Registro_patronal')))
        
        # Receptor nómina
        receptor_nom = ET.SubElement(nomina, '{http://www.sat.gob.mx/nomina12}Receptor')
        receptor_nom.set('Curp', self.limpiar_texto(empleado.get('CURP', '')))
        ##mati
        receptor_nom.set('Antigüedad',  self.antiguedad_pac(empleado.get('Fecha_inicio_relacion_laboral'), empleado.get('Fecha_final_pago')))
        receptor_nom.set('RiesgoPuesto', '1') ##FALTA AGREGAR RIESGO DE PUESTO REAL, POR AHORA SE PONE 1 (Riesgo mínimo) PARA EVITAR RECHAZOS DEL PAC
        #
        # Campos opcionales del receptor - solo si tienen valor y no están excluidos
        campos_opcionales = {
            'NumSeguridadSocial': 'No_Seguro_social',
            'FechaInicioRelLaboral': 'Fecha_inicio_relacion_laboral',
            'Departamento': 'Departamento',
            'Puesto': 'Puesto',
            #'Banco': 'Banco', ###mati
            #'CuentaBancaria': 'Cuenta bancaria',
            'SalarioBaseCotApor': 'Salario_base_cot_apor',
            'SalarioDiarioIntegrado': 'Salario_diario_integrado'
        }
        
        for attr_xml, campo_excel in campos_opcionales.items():
            if not ConfiguracionCampos.campo_excluido(attr_xml):
                valor = empleado.get(campo_excel)
                if self.tiene_valor(valor):
                    # Formatear fechas si es necesario
                    if 'Fecha' in attr_xml:
                        valor = self.formatear_fecha_simple(valor)
                    # Formatear números si es necesario
                    elif attr_xml in ['SalarioBaseCotApor', 'SalarioDiarioIntegrado']:
                        valor = f"{float(valor):.2f}"
                    # Limpiar CuentaBancaria de notación científica
                    elif attr_xml == 'CuentaBancaria':
                        try:
                            # Convertir de notación científica a entero
                            valor = str(int(float(valor)))
                        except:
                            valor = str(valor)
                    
                    # Aplicar limpieza de texto para todos los campos de texto
                    if isinstance(valor, str):
                        valor = self.limpiar_texto(valor)
                    
                    receptor_nom.set(attr_xml, str(valor))
        
        # Campos obligatorios - aplicar limpieza
        if self.tiene_valor(empleado.get('Tipo_contrato')):
            tipo_contrato = self.limpiar_texto(empleado.get('Tipo_contrato')).split('-')[0].strip()
            receptor_nom.set('TipoContrato', tipo_contrato)
        
        # Sindicalizado
        if self.tiene_valor(empleado.get('Sindicalizado')):
            receptor_nom.set('Sindicalizado', self.limpiar_texto(empleado.get('Sindicalizado')))
        
        # TipoJornada es OBLIGATORIO según XSD
        if self.tiene_valor(empleado.get('Tipo jornada')):
            tipo_jornada = self.limpiar_texto(empleado.get('Tipo jornada')).split('-')[0].strip()
            receptor_nom.set('TipoJornada', tipo_jornada)
        else:
            receptor_nom.set('TipoJornada', '01')  # Default: Diurna
        
        if self.tiene_valor(empleado.get('Tipo_regimen')):
            tipo_regimen = self.limpiar_texto(empleado.get('Tipo_regimen')).split('-')[0].strip()
            receptor_nom.set('TipoRegimen', tipo_regimen)
        if self.tiene_valor(empleado.get('Num_empleado')):
            receptor_nom.set('NumEmpleado', self.limpiar_texto(empleado.get('Num_empleado')))
        if self.tiene_valor(empleado.get('Periodicidad_pago')):
            receptor_nom.set('PeriodicidadPago', self.ceor_izquierda(self.limpiar_texto(empleado.get('Periodicidad_pago'))))
        if self.tiene_valor(empleado.get('ClaveEntFed')):
            receptor_nom.set('ClaveEntFed', self.limpiar_texto(empleado.get('ClaveEntFed')))
        
        # Percepciones - solo si hay percepciones con valor
        if percepciones:
            percepciones_elem = ET.SubElement(nomina, '{http://www.sat.gob.mx/nomina12}Percepciones')
            percepciones_elem.set('TotalSueldos', f"{float(totales_calc['total_sueldos']):.2f}")
            percepciones_elem.set('TotalGravado', f"{float(totales_calc['total_gravado']):.2f}")
            percepciones_elem.set('TotalExento', f"{float(totales_calc['total_exento']):.2f}")
            
            for perc in percepciones:
                # Solo agregar si al menos uno de los importes es mayor a 0
                if perc['gravado'] > 0 or perc['exento'] > 0:
                    perc_elem = ET.SubElement(percepciones_elem, '{http://www.sat.gob.mx/nomina12}Percepcion')
                    perc_elem.set('TipoPercepcion', perc['tipo'])
                    perc_elem.set('Clave', perc['clave'])  # CAMBIADO: usar la clave del Excel
                    perc_elem.set('Concepto', self.limpiar_texto(perc['concepto']))
                    perc_elem.set('ImporteGravado', f"{float(perc['gravado']):.2f}")
                    perc_elem.set('ImporteExento', f"{float(perc['exento']):.2f}")
                    
                    # Si es tipo 019 (Horas Extra), agregar nodo HorasExtra
                    if perc['tipo'] == '019':
                        horas_extra = ET.SubElement(perc_elem, '{http://www.sat.gob.mx/nomina12}HorasExtra')
                        # Valores por defecto - el usuario debe ajustarlos según sus datos
                        horas_extra.set('Dias', str(empleado.get('Dias', '1')))
                        horas_extra.set('TipoHoras', str(empleado.get('TipoHoras', '01')))  # 01=Dobles
                        horas_extra.set('HorasExtra', str(empleado.get('HorasExtra', '2')))
                        horas_extra.set('ImportePagado', f"{float(perc['gravado'] + perc['exento']):.2f}")
        
        # Deducciones - solo si hay deducciones con valor
        if deducciones:
            deducciones_elem = ET.SubElement(nomina, '{http://www.sat.gob.mx/nomina12}Deducciones')
            deducciones_elem.set('TotalOtrasDeducciones', f"{float(totales_calc['total_otras_deducciones']):.2f}")
            deducciones_elem.set('TotalImpuestosRetenidos', f"{float(totales_calc['total_impuestos_retenidos']):.2f}")
            
            for ded in deducciones:
                ded_elem = ET.SubElement(deducciones_elem, '{http://www.sat.gob.mx/nomina12}Deduccion')
                ded_elem.set('TipoDeduccion', ded['tipo'])
                ded_elem.set('Clave', ded['clave'])  # CAMBIADO: usar la clave del Excel
                ded_elem.set('Concepto', self.limpiar_texto(ded['concepto']))
                ded_elem.set('Importe', f"{float(ded['importe']):.2f}")
        
        # Otros pagos - solo si hay otros pagos con valor > 0
        if otros_pagos:
            otros_pagos_elem = ET.SubElement(nomina, '{http://www.sat.gob.mx/nomina12}OtrosPagos')
            
            for op in otros_pagos:
                otro_elem = ET.SubElement(otros_pagos_elem, '{http://www.sat.gob.mx/nomina12}OtroPago')
                otro_elem.set('TipoOtroPago', op['tipo'])
                otro_elem.set('Clave', op['clave'])  # CAMBIADO: usar la clave del Excel
                otro_elem.set('Concepto', self.limpiar_texto(op['concepto']))
                otro_elem.set('Importe', f"{float(op['importe']):.2f}")
                
                # Si es subsidio al empleo (002), agregar nodo SubsidioAlEmpleo
                if op['tipo'] == '002':
                    subsidio_elem = ET.SubElement(otro_elem, '{http://www.sat.gob.mx/nomina12}SubsidioAlEmpleo')
                    subsidio_elem.set('SubsidioCausado', f"{float(op['importe']):.2f}")
        
        return comprobante
    
    def crear_reporte_diferencias(self, empleado, totales_calc):
        """Crea un reporte de diferencias entre valores del Excel y calculados"""
        reporte = {
            'empleado': empleado.get('Nombre', 'Sin nombre'),
            'rfc': empleado.get('RFC_receptor', 'Sin RFC'),
            'diferencias': []
        }
        
        # Comparar totales - solo si el campo del Excel tiene valor
        campos_comparar = [
            ('Subtotal', 'subtotal'),
            ('Total', 'total'),
            ('Total_percepciones', 'total_percepciones'),
            ('Total_deducciones', 'total_deducciones'),
            ('Total_otros_pagos', 'total_otros_pagos'),
            ('Total_gravado', 'total_gravado'),
            ('TotalExento', 'total_exento'),
            ('Total_sueldos', 'total_sueldos'),
            ('TotalOtrasDeducciones', 'total_otras_deducciones'),
            ('TotalImpuestosRetenidos', 'total_impuestos_retenidos')
        ]
        
        for campo_excel, campo_calc in campos_comparar:
            valor_excel = empleado.get(campo_excel)
            valor_calc = totales_calc.get(campo_calc)
            
            # Solo comparar si el valor del Excel tiene valor (no es None)
            if self.tiene_valor(valor_excel) and valor_calc is not None:
                valor_excel_dec = Decimal(str(valor_excel))
                diferencia = abs(valor_excel_dec - valor_calc)
                
                if diferencia > Decimal('0.01'):  # Diferencia mayor a 1 centavo
                    reporte['diferencias'].append({
                        'campo': campo_excel,
                        'valor_excel': float(valor_excel_dec),
                        'valor_calculado': float(valor_calc),
                        'diferencia': float(diferencia)
                    })
        
        return reporte
    
    def formatear_xml(self, elem, level=0):
        """Formatea el XML con indentación"""
        indent = "    "
        i = "\n" + level * indent
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + indent
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for child in elem:
                self.formatear_xml(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i
    
    def generar_nombre_archivo(self, empleado):
        """Genera el nombre del archivo XML"""
        rfc = empleado.get('RFC_receptor', 'SINRFC')
        folio = empleado.get('Folio', '0000')
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
        uid = uuid.uuid4().hex[:8]  # solo 8 caracteres para no hacerlo muy largo

        return f"CFDI_{rfc}_{folio}_{timestamp}_{uid}.xml".upper()
    
    def generar_reporte_general(self):
        """Genera un reporte general de todos los XMLs procesados"""
        reporte_lines = []
        reporte_lines.append("=" * 80)
        reporte_lines.append("REPORTE DE GENERACIÓN DE XMLs DE NÓMINA DESDE EXCEL")
        reporte_lines.append("=" * 80)
        reporte_lines.append("")
        
        # Información de configuración
        reporte_lines.append("CONFIGURACIÓN DE CAMPOS EXCLUIDOS:")
        reporte_lines.append("-" * 80)
        if ConfiguracionCampos.CAMPOS_EXCLUIDOS_RECEPTOR_NOMINA:
            reporte_lines.append("  Campos excluidos del Receptor de Nómina:")
            for campo in ConfiguracionCampos.CAMPOS_EXCLUIDOS_RECEPTOR_NOMINA:
                reporte_lines.append(f"    • {campo}")
        else:
            reporte_lines.append("  No hay campos excluidos del Receptor de Nómina")
        
        reporte_lines.append("")
        
        for idx, reporte in enumerate(self.reportes, 1):
            reporte_lines.append(f"\n{idx}. EMPLEADO: {reporte['empleado']}")
            reporte_lines.append(f"   RFC: {reporte['rfc']}")
            reporte_lines.append("-" * 80)
            
            if reporte['diferencias']:
                reporte_lines.append("   DIFERENCIAS ENCONTRADAS:")
                for dif in reporte['diferencias']:
                    reporte_lines.append(f"      • {dif['campo']}:")
                    reporte_lines.append(f"        - Excel:      ${dif['valor_excel']:,.2f}")
                    reporte_lines.append(f"        - Calculado:  ${dif['valor_calculado']:,.2f}")
                    reporte_lines.append(f"        - Diferencia: ${dif['diferencia']:,.2f}")
            else:
                reporte_lines.append("   ✓ No se encontraron diferencias")
        
        reporte_lines.append("\n" + "=" * 80)
        reporte_lines.append(f"Total de empleados procesados: {len(self.reportes)}")
        reporte_lines.append("=" * 80)
        
        return "\n".join(reporte_lines)
    
    def procesar(self):
        """Procesa el Excel y genera los XMLs"""
        print("Iniciando procesamiento de Excel...")
        
        # Crear directorio de salida si no existe
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Guardar configuración actual
        config_path = os.path.join(self.output_dir, 'configuracion_campos.json')
        ConfiguracionCampos.guardar_configuracion(config_path)
        print(f"   ✓ Configuración guardada en: {config_path}")
        
        # Leer Excel
        print("1. Leyendo archivo Excel...")
        empleados = self.leer_excel()
        print(f"   ✓ {len(empleados)} empleado(s) encontrado(s)")
        archivos_procesados = []
        # Generar XML por cada empleado
        print("2. Generando XMLs...")
        ok, fail = 0, 0
        with conexion() as db:
            for idx, empleado in enumerate(empleados, 1):
                nombre_archivo = self.generar_nombre_archivo(empleado)
                ruta_completa = os.path.join(self.output_dir, empleado.get('RFC_emisor', 'default'), 'CFDI', nombre_archivo)
                os.makedirs(os.path.dirname(ruta_completa), exist_ok=True)
                excel_row = empleado.get('__excel_row__', 'N/A')
                datos_insert ={
                    'rfc_receptor': empleado.get('RFC_receptor'),
                    'rfc_emisor': empleado.get('RFC_emisor'),
                    'folio': empleado.get('Folio'),
                    'fecha_generado': datetime.now(),
                    'fila_excel': excel_row,
                    'nombre_archivo': nombre_archivo,
                    'ruta_xml': ruta_completa,
                    'generado': True,
                    "archivo_origen": self.excel_path,
                    "ultima_actualizacion": datetime.now(),
                    'ambiente': self.ambiente
                }
                # Generar XML
                try:
                    doc = self.generar_xml(empleado)
                    # Formatear y guardar
                    self.formatear_xml(doc)
                    tree = ET.ElementTree(doc)
                    
                    tree.write(ruta_completa, encoding='utf-8', xml_declaration=True, method='xml')
                    if(os.path.exists(ruta_completa)):
                        archivos_procesados.append(ruta_completa)
                        db.insert(datos_insert, 'nomina.procesos_cfdi')
                        ok += 1
                except Exception as e:
                    # Mensaje con fila exacta y contexto
                    err = f"Falla en fila Excel {excel_row}. Error: {traceback.format_exc()}"
                    datos_insert['ruta_xml'] = None
                    datos_insert['nombre_archivo'] = f"ERROR_FILA_{excel_row}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xml"
                    datos_insert['generado'] = False
                    datos_insert['error'] = err
                    db.insert(datos_insert, 'nomina.procesos_cfdi')
                    fail += 1

        print(f"   ✓ XMLs generados exitosamente: {ok} | Fallidos: {fail}")
        # Generar reporte
        print("3. Generando reporte...")
        reporte = self.generar_reporte_general()
        ruta_reporte = os.path.join(self.output_dir, 'reporte_generacion.txt')
        with open(ruta_reporte, 'w', encoding='utf-8') as f:
            f.write(reporte)
        return archivos_procesados
            

if __name__ == "__main__":
    procesador = ExcelToNominaXML(
        excel_path='tmp/TEMPLATE_2.xlsx',
        output_dir='tmp/XMLs'
    )
    
    print(procesador.procesar())